import datetime
import os
import sqlite3
from django.conf import settings
from django.core.cache import cache
from django.db import connection, OperationalError
from django.db.models import CharField, TextField, ForeignKey, IntegerField, AutoField
import pandas as pd
from core.models import Item
import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from djmoney.models.fields import MoneyFieldProxy

# Adapted from https://stackoverflow.com/a/48457168
def trunc_datetime(date:datetime.datetime):
    """Zeroes out the hour, minute, second, and microsecond elements of a datetime object for ease of comparison of dates alone.

    Args:
        date (datetime.datetime): The datetime to truncate.

    Returns:
        datetime.datetime: The truncated datetime.
    """
    return date.replace(hour=0, minute=0, second=0, microsecond=0)

def dict_from_excel_row(row: pd.Series) -> dict:
    """Creates a dictionary from an Excel row, corresponding to the predefined models used in this Django project.

    Args:
        row (pd.Series): _description_

    Raises:
        KeyError: _description_
        Exception: _description_
        KeyError: _description_
        KeyError: _description_

    Returns:
        dict: _description_
    """
    data = {}
    try:
        # Identify item by ITEM_NO or ITEM
        item_no = row.get("ITEM_NO")
        item_name = row.get("ITEM")
        item_identifier = item_no or item_name

        if not item_identifier:
            raise KeyError("Missing ITEM or ITEM_NO for Item lookup or creation.")

        # Prepare defaults for a new Item if needed
        item_defaults = {
            "item": item_name or "",
            "item_no": item_no or "",
            "mfr": row.get("MFR", ""),
            "mfr_cat": row.get("MFR CAT", ""),
            "descr": row.get("DESCR", ""),
            "par_level": 1,
            "external_url": "https://accessgudid.nlm.nih.gov/resources/developers/v3/device_lookup_api",
        }

        item_instance, _ = Item.objects.get_or_create(item_no=item_identifier, defaults=item_defaults)
        data["item"] = item_instance

        if "VENDOR" in row:
            data["vendor"] = row["VENDOR"]
        if "VEND CAT" in row:
            data["vend_cat"] = row["VEND CAT"]
        if "RECV QTY" in row:
            data["recv_qty"] = row["RECV QTY"]
        if "UM" in row:
            data["um"] = row["UM"]
        if "PRICE" in row:
            data["price"] = row["PRICE"]
        if "TOTAL COST" in row:
            data["total_cost"] = row["TOTAL COST"]
        if "Expr1010" in row:
            data["expr1010"] = row["Expr1010"]
        if "PO_NO" in row:
            data["po_no"] = row["PO_NO"]
        if "PO_DATE" in row:
            data["po_date"] = row["PO_DATE"].tz_localize(tz="America/Chicago").tz_convert("UTC")
        if "VEND_CODE" in row:
            data["vend_code"] = row["VEND_CODE"]
        if "dbo_VEND.NAME" in row:
            data["dbo_vend_name"] = row["dbo_VEND.NAME"]
    except KeyError as e:
        raise Exception("Missing required field value") from e

    if "dbo_CC.NAME" in row:
        data["dbo_cc_name"] = row["dbo_CC.NAME"]
    elif "Expr1016" in row:
        data["dbo_cc_name"] = row["Expr1016"]
    else:
        raise KeyError('Key either "dbo_CC.NAME" or "Expr1016" required.')

    if "ACCT_NO" in row:
        data["acct_no"] = row["ACCT_NO"]
    elif "Expr1017" in row:
        data["acct_no"] = row["Expr1017"]
    else:
        raise KeyError('Key either "ACCT_NO" or "Expr1017" required.')

    if "RCV_DATE" in row:
        data["rcv_date"] = row["RCV_DATE"].tz_localize(tz="America/Chicago").tz_convert("UTC")

    return data

def style_excel_sheet(sheet, type, field, i, currency_style):
    """Styles the exported Excel sheet to match the original style received to use for importing.

    Args:
        sheet (_type_): The openpyxl active sheet
        type (_type_): The model type by which to check field names against
        field (_type_): The current field to format the corresponding Excel column for.
        i (_type_): The current column index of the sheet
        currency_style (_type_): The openpyxl named style to use for formatting currency.
    """
    col_letter = openpyxl.utils.get_column_letter(i)
    if isinstance(getattr(type, field.name), MoneyFieldProxy):
        for row in range(2, sheet.max_row + 1):  # Start from the second row (skip header)
            sheet[f'{col_letter}{row}'].style = currency_style  # Apply the style to the entire column
    
    # apply header formatting
    cell_to_format = sheet.cell(row=1, column=i)
    cell_to_format.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    cell_to_format.font = Font(bold=True, color="000000")
    cell_to_format.alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    cell_to_format.border = thin_border


    max_length = 0
    for cell in list(sheet.columns)[i-1]:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    
    sheet.column_dimensions[col_letter].width = (max_length + 2)

def absolute_add_remove_quantity(item_quantity: int, add_remove_mode: str):
    quantity = abs(item_quantity)
    if add_remove_mode.lower() == "out":
        quantity = quantity * -1
    
    return quantity

def get_searchable_fields(model):
    fields = []
    for f in model._meta.get_fields():
        # Direct text fields
        if isinstance(f, (CharField, TextField)) and not f.is_relation:
            fields.append(f.name)

        # Direct numeric fields
        elif isinstance(f, (IntegerField, AutoField)) and not f.is_relation:
            fields.append(f.name)

        # ForeignKeys -> dive into related model for text fields
        elif isinstance(f, ForeignKey):
            related_model = f.related_model
            for rf in related_model._meta.get_fields():
                if isinstance(rf, (CharField, TextField)) and not rf.is_relation:
                    fields.append(f"{f.name}__{rf.name}")
    
    return fields

def get_database_status():
    status = cache.get('database_status')
    if status is None:
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT 1')
                cursor.fetchone()
            db_file = settings.DATABASES['default']['NAME']
            if not os.path.exists(db_file):
                status = "Disconnected: Database file not found"
            elif not os.access(db_file, os.R_OK | os.W_OK):
                status = "Disconnected: Insufficient permissions for database file"
            else:
                with sqlite3.connect(db_file) as conn:
                    cursor = conn.cursor()
                    cursor.execute('PRAGMA integrity_check')
                    result = cursor.fetchone()
                    status = "Connected" if result[0] == 'ok' else f"Disconnected: Database integrity check failed ({result[0]})"
        except OperationalError as e:
            status = f"Disconnected: Database error ({str(e)})"
        except Exception as e:
            status = f"Disconnected: Unexpected error ({str(e)})"
        cache.set('database_status', status, timeout=60)  # Cache for 60 seconds
    return status